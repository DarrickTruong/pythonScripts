import openpyxl
import re
import os

os.chdir('/Users/DarrickT/Downloads/')
wb = openpyxl.load_workbook('surfraw.xlsx')
sheet = wb['Page 1']
def oldestDate():
    for i in range(2, 8657):
        print(sheet.cell(row=i,column=26).value)
        all1 = re.findall(r"[\d]{4}-[\d]{2}-[\d]{2}", sheet.cell(row=i,column=26).value)
        all2 = re.findall(r"[\d]{4}-[\d]{2}-[\d]{2}", sheet.cell(row=i,column=27).value)
        if len(all1) !=0:
            sheet.cell(row=i, column=52).value = all1[len(all1)-1]
        if len(all2) !=0:
            sheet.cell(row=i, column=53).value = all2[len(all2)-1]
    wb.save('surfraw1.xlsx')

