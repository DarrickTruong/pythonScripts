import os
import re

import openpyxl

os.chdir("/Users/darrick/Downloads/")

def run(numSensors, numRows, fileName):
    wb = openpyxl.load_workbook(fileName)
    print(type(wb))
    sheet = wb["Sheet1"]
    new = 1
    for i in range(1,numRows, numSensors+1):
        # # print(type(str(sheet.cell(row=i, column=1).value)))
        # # x = re.search("[0-9]+", sheet.cell(row=i, column=1).value)
        # sheet.cell(row=j, column=4).value = sheet.cell(
        #     row=i, column=1).value
        # sheet.cell(row=j, column=3).value = sheet.cell(
        #     row=i+1, column=1).value
        # j+=1
        
        
        # if isinstance(sheet.cell(row=i, column=1).value, int):
        #     print("in x")
        #     sheet.cell(row=i, column=4).value = sheet.cell(row=i, column=1).value
        # else:
        #     print("in else")
        #     sheet.cell(row=i-1, column=3).value = sheet.cell(row=i, column=1).value
        for j in range(0, numSensors):
            sheet.cell(row=new+j, column=4).value = sheet.cell(row=i+j, column=1).value
            # sheet.cell(row=j+1, column=4).value = sheet.cell(row=i+1, column=1).value
            # sheet.cell(row=j+2, column=4).value = sheet.cell(row=i+2, column=1).value
            # sheet.cell(row=j+3, column=4).value = sheet.cell(row=i+3, column=1).value
        sheet.cell(row=new, column=3).value = sheet.cell(row=i+numSensors, column=1).value
        new+= numSensors

    wb.save('idle3.xlsx')
